"""
Excel Export Generator with Styling
Creates a professionally styled Excel report from validation results.
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


class ExcelExporter:
    def __init__(self, results_file: str = "results/validation_results.json"):
        self.results_file = Path(results_file)
        self.results = self._load_results()
        
        # Define color schemes
        self.colors = {
            'header': 'FFE082FF',      # Purple header
            'success': 'FFC6EFCE',     # Light green
            'warning': 'FFFFFF9C',     # Light yellow
            'error': 'FFFFC7CE',       # Light red
            'no_schema': 'FFFFCC99',   # Light orange
            'blocked': 'FFE0E0E0',     # Light gray
            'alt_row': 'FFF5F5F5',     # Alternating row
        }
    
    def _load_results(self):
        """Load validation results from JSON file."""
        if not self.results_file.exists():
            return []
        
        with open(self.results_file, 'r') as f:
            return json.load(f)
    
    def generate_excel(self, output_file: str = "results/validation_report.xlsx"):
        """Generate styled Excel report."""
        output_path = Path(output_file)
        output_path.parent.mkdir(exist_ok=True)
        
        wb = Workbook()
        
        # Create sheets
        ws_full = wb.active
        ws_full.title = "Full Results"
        
        # Generate full results sheet
        self._create_full_results_sheet(ws_full)
        
        # Create summary sheet
        ws_summary = wb.create_sheet("Summary", 0)
        self._create_summary_sheet(ws_summary)
        
        # Save workbook
        wb.save(output_path)
        return output_path
    
    def _create_summary_sheet(self, ws):
        """Create summary dashboard sheet."""
        # Title
        ws['A1'] = 'Product Schema Validation Report'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='6A0DAD', end_color='6A0DAD', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        # Generation time
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A2'].font = Font(italic=True, color='666666')
        ws.merge_cells('A2:D2')
        
        # Summary stats
        total = len(self.results)
        success = sum(1 for r in self.results if r['status'] == 'success')
        warning = sum(1 for r in self.results if r['status'] == 'warning')
        error = sum(1 for r in self.results if r['status'] == 'error')
        no_schema = sum(1 for r in self.results if r['status'] == 'no_schema')
        blocked = sum(1 for r in self.results if r['status'] == 'blocked')
        schema_found = sum(1 for r in self.results if r.get('schema_found', False))
        
        # Stats table
        row = 4
        headers = ['Metric', 'Count', 'Percentage', '']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        stats = [
            ('Total URLs', total, 100),
            ('Success', success, round(success/total*100, 1) if total > 0 else 0),
            ('Warning', warning, round(warning/total*100, 1) if total > 0 else 0),
            ('Error', error, round(error/total*100, 1) if total > 0 else 0),
            ('No Schema', no_schema, round(no_schema/total*100, 1) if total > 0 else 0),
            ('Blocked', blocked, round(blocked/total*100, 1) if total > 0 else 0),
            ('Schema Found', schema_found, round(schema_found/total*100, 1) if total > 0 else 0),
        ]
        
        for idx, (metric, count, pct) in enumerate(stats, row + 1):
            ws.cell(idx, 1, metric).font = Font(bold=True)
            ws.cell(idx, 2, count).alignment = Alignment(horizontal='center')
            ws.cell(idx, 3, f'{pct}%').alignment = Alignment(horizontal='center')
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
    
    def _create_full_results_sheet(self, ws):
        """Create full results sheet with all data."""
        # Headers
        headers = ['URL', 'Status', 'Schema Found', 'Score', '# Errors', '# Warnings', 'Errors', 'Warnings', 'Response Time']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                bottom=Side(style='medium', color='000000')
            )
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add data
        for row_idx, result in enumerate(self.results, 2):
            # Determine status color
            status = result.get('status', 'unknown')
            status_color = self.colors.get(status, 'FFFFFF')
            
            # Alternating row background for readability
            row_fill = PatternFill(start_color=self.colors['alt_row'], end_color=self.colors['alt_row'], fill_type='solid') if row_idx % 2 == 0 else None
            
            # URL (with hyperlink)
            url_cell = ws.cell(row_idx, 1, result['url'])
            url_cell.hyperlink = result['url']
            url_cell.font = Font(color='0000FF', underline='single')
            if row_fill:
                url_cell.fill = row_fill
            
            # Status (color-coded)
            status_cell = ws.cell(row_idx, 2, status.upper())
            status_cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type='solid')
            status_cell.alignment = Alignment(horizontal='center')
            status_cell.font = Font(bold=True)
            
            # Schema Found
            schema_cell = ws.cell(row_idx, 3, 'Yes' if result.get('schema_found') else 'No')
            schema_cell.alignment = Alignment(horizontal='center')
            if row_fill:
                schema_cell.fill = row_fill
            
            # Score
            score = 0
            validation = result.get('validation')
            if validation and validation.get('score') is not None:
                score = validation['score']
            elif result.get('score') is not None:
                score = result['score']
            
            score_cell = ws.cell(row_idx, 4, f'{score}%')
            score_cell.alignment = Alignment(horizontal='center')
            # Color gradient based on score
            if score >= 90:
                score_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            elif score >= 70:
                score_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            elif score > 0:
                score_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            elif row_fill:
                score_cell.fill = row_fill
            
            # Error Count
            error_count = 0
            errors_list = []
            if validation and validation.get('errors'):
                error_count = len(validation['errors'])
                errors_list = validation['errors']
            elif result.get('errors') and isinstance(result['errors'], list):
                error_count = len(result['errors'])
                errors_list = result['errors']
            elif result.get('error'):
                error_count = 1
                errors_list = [result['error']]
            
            error_count_cell = ws.cell(row_idx, 5, error_count)
            error_count_cell.alignment = Alignment(horizontal='center')
            if error_count > 0:
                error_count_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                error_count_cell.font = Font(bold=True, color='9C0006')
            elif row_fill:
                error_count_cell.fill = row_fill
            
            # Warning Count
            warning_count = 0
            warnings_list = []
            if validation and validation.get('warnings'):
                warning_count = len(validation['warnings'])
                warnings_list = validation['warnings']
            elif result.get('warnings') and isinstance(result['warnings'], list):
                warning_count = len(result['warnings'])
                warnings_list = result['warnings']
            
            warning_count_cell = ws.cell(row_idx, 6, warning_count)
            warning_count_cell.alignment = Alignment(horizontal='center')
            if warning_count > 0:
                warning_count_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                warning_count_cell.font = Font(bold=True, color='9C5700')
            elif row_fill:
                warning_count_cell.fill = row_fill
            
            # Errors (detailed)
            errors_text = '; '.join([
                e if isinstance(e, str) else e.get('message', '') 
                for e in errors_list
            ])
            errors_cell = ws.cell(row_idx, 7, errors_text)
            errors_cell.alignment = Alignment(wrap_text=True, vertical='top')
            if row_fill:
                errors_cell.fill = row_fill
            
            # Warnings (detailed)
            warnings_text = '; '.join([
                w if isinstance(w, str) else w.get('message', '') 
                for w in warnings_list
            ])
            warnings_cell = ws.cell(row_idx, 8, warnings_text)
            warnings_cell.alignment = Alignment(wrap_text=True, vertical='top')
            if row_fill:
                warnings_cell.fill = row_fill
            
            # Response Time
            response_time = result.get('response_time', 0)
            time_cell = ws.cell(row_idx, 9, f'{response_time:.2f}s' if response_time else 'N/A')
            time_cell.alignment = Alignment(horizontal='center')
            if row_fill:
                time_cell.fill = row_fill
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 50  # URL
        ws.column_dimensions['B'].width = 12  # Status
        ws.column_dimensions['C'].width = 13  # Schema Found
        ws.column_dimensions['D'].width = 10  # Score
        ws.column_dimensions['E'].width = 12  # # Errors
        ws.column_dimensions['F'].width = 13  # # Warnings
        ws.column_dimensions['G'].width = 50  # Errors
        ws.column_dimensions['H'].width = 50  # Warnings
        ws.column_dimensions['I'].width = 15  # Response Time
        
        # Add auto-filter
        ws.auto_filter.ref = f'A1:I{len(self.results) + 1}'


def main():
    """Generate Excel report."""
    exporter = ExcelExporter()
    output = exporter.generate_excel()
    print(f'✅ Excel report generated: {output}')


if __name__ == '__main__':
    main()

